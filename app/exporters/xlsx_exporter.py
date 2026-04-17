from __future__ import annotations

from pathlib import Path
from typing import Any

from app.core.models import (
    AGGREGATES_COLUMNS,
    EXTERNAL_LINKS_COLUMNS,
    HIGHLIGHTS_COLUMNS,
    POSTS_COLUMNS,
    PROFILE_COLUMNS,
    RUN_LOG_COLUMNS,
)

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None
    get_column_letter = None


def _write_sheet(
    workbook: Any, name: str, columns: list[str], rows: list[dict[str, Any]]
) -> None:
    ws = workbook.create_sheet(title=name)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col) for col in columns])
    _autofit_columns(ws)


def _autofit_columns(worksheet: Any, max_width: int = 64) -> None:
    if get_column_letter is None:
        return
    for index, column_cells in enumerate(worksheet.iter_cols(), start=1):
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        width = min(max(max_len + 2, 10), max_width)
        worksheet.column_dimensions[get_column_letter(index)].width = width


def export_xlsx_artifacts(
    exports_dir: Path,
    base_name: str,
    run_log_rows: list[dict[str, Any]],
    profile_rows: list[dict[str, Any]],
    highlights_rows: list[dict[str, Any]],
    external_links_rows: list[dict[str, Any]],
    posts_rows: list[dict[str, Any]],
    aggregate_rows: list[dict[str, Any]],
    summary_flat_rows: list[dict[str, Any]],
) -> dict[str, str]:
    if Workbook is None:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for XLSX export. Install with: pip install openpyxl"
        )

    exports_dir.mkdir(parents=True, exist_ok=True)
    normalized_path = exports_dir / f"{base_name}_normalized.xlsx"
    summary_path = exports_dir / f"{base_name}_master_summary.xlsx"

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    _write_sheet(wb, "run_log", RUN_LOG_COLUMNS, run_log_rows)
    _write_sheet(wb, "profile", PROFILE_COLUMNS, profile_rows)
    _write_sheet(wb, "highlights", HIGHLIGHTS_COLUMNS, highlights_rows)
    _write_sheet(wb, "external_links", EXTERNAL_LINKS_COLUMNS, external_links_rows)
    _write_sheet(wb, "posts", POSTS_COLUMNS, posts_rows)
    _write_sheet(wb, "aggregates", AGGREGATES_COLUMNS, aggregate_rows)
    wb.save(normalized_path)

    wb2 = Workbook()
    ws = wb2.active
    ws.title = "master_summary"
    columns = (
        list(summary_flat_rows[0].keys()) if summary_flat_rows else ["scraped_at_ist"]
    )
    ws.append(columns)
    for row in summary_flat_rows:
        ws.append([row.get(c) for c in columns])
    _autofit_columns(ws)
    wb2.save(summary_path)

    return {
        "normalized_xlsx": str(normalized_path.resolve()),
        "summary_flat_xlsx": str(summary_path.resolve()),
        "master_summary_xlsx": str(summary_path.resolve()),
    }
